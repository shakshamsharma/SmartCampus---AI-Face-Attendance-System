import io
import csv
import json
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_attendance_pdf(title: str, data: list, summary: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1a1a2e'), spaceAfter=4)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666'), spaceAfter=12)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#4f8ef7'), spaceBefore=16, spaceAfter=8, fontName='Helvetica-Bold')
    
    story = []
    
    # Header
    story.append(Paragraph("🎓 SmartCampus AI", title_style))
    story.append(Paragraph("AI-Powered Attendance & Analytics Platform", sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#4f8ef7')))
    story.append(Spacer(1, 12))
    story.append(Paragraph(title, ParagraphStyle('RepTitle', parent=styles['Normal'], fontSize=14, fontName='Helvetica-Bold', spaceAfter=4)))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", sub_style))
    story.append(Spacer(1, 16))
    
    # Summary cards
    if summary:
        story.append(Paragraph("📊 Summary", header_style))
        sum_data = [list(summary.keys()), list(summary.values())]
        sum_table = Table([sum_data[0], sum_data[1]], colWidths=[120]*len(summary))
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f0f4ff')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#4f8ef7')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9ff')]),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#e0e8ff')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e0e8ff')),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(sum_table)
        story.append(Spacer(1, 16))
    
    # Main data table
    if data:
        story.append(Paragraph("📋 Attendance Details", header_style))
        headers = list(data[0].keys())
        rows = [headers] + [[str(row.get(h, "")) for h in headers] for row in data]
        col_width = (A4[0] - 80) / len(headers)
        table = Table(rows, colWidths=[col_width]*len(headers), repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f8ef7')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f7ff')]),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#c0d0ff')),
            ('INNERGRID', (0,0), (-1,-1), 0.3, colors.HexColor('#dde8ff')),
            ('TOPPADDING', (0,0), (-1,-1), 7),
            ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ]))
        story.append(table)
    
    story.append(Spacer(1, 24))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#ddd')))
    story.append(Paragraph("Confidential — SmartCampus AI Platform | AI-Powered Attendance Intelligence", 
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#999'), alignment=TA_CENTER, spaceBefore=8)))
    
    doc.build(story)
    buf.seek(0)
    return buf.read()

def generate_attendance_csv(data: list) -> bytes:
    if not data:
        return b"No data available"
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    return buf.getvalue().encode()

def generate_attendance_excel(title: str, data: list, summary: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Header
    ws['A1'] = "SmartCampus AI — Attendance Report"
    ws['A1'].font = Font(size=16, bold=True, color="1a1a2e")
    ws['A2'] = title
    ws['A2'].font = Font(size=12, color="4f8ef7")
    ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws['A3'].font = Font(size=10, color="666666")
    ws.merge_cells('A1:F1')
    ws.merge_cells('A2:F2')
    ws.merge_cells('A3:F3')
    
    row = 5
    
    # Summary
    if summary:
        ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, color="4f8ef7", size=11)
        row += 1
        for k, v in summary.items():
            ws.cell(row=row, column=1, value=k).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(v))
            row += 1
        row += 1
    
    # Data headers
    if data:
        headers = list(data[0].keys())
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4f8ef7")
            cell.alignment = Alignment(horizontal="center")
        row += 1
        
        for record in data:
            for col_idx, h in enumerate(headers, 1):
                val = record.get(h, "")
                cell = ws.cell(row=row, column=col_idx, value=str(val))
                cell.alignment = Alignment(horizontal="center")
            row += 1
        
        for col_idx in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
